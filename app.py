import streamlit as st
from openpyxl import load_workbook
import tempfile
import io
import re
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

# Configuración de la página
st.set_page_config(page_title="🍽️ Adaptador de Menús con PDF")

st.title("🍽️ Adaptador de Menús")
st.write("Sube tu menú en Excel (.xlsx). Esta app cambiará alimentos concretos y te permite descargar el archivo adaptado en Excel o PDF.")

uploaded_file = st.file_uploader("📤 Sube tu archivo Excel", type=["xlsx"])

# Diccionario de sustituciones
REEMPLAZOS = {
    "salchichas frescas": "Pechuga de pavo al horno",
    "croquetas de jamón": "Filete de aguja en su jugo",
    "ensalada césar (lechuga, pollo, queso y salsa césar)": "Ensalada variada con pollo",
    "olleta alicantina": "Legumbres (no lentejas)",
    "pizza margarita": "Pizza sin gluten",
    "albóndigas con salsa": "Albóndigas sin gluten",
    "buñuelos de bacalao": "Buñuelos sin gluten (maicena)",
    "lomo adobado": "Lomo fresco",
}

def adaptar_valor(valor):
    if valor is None or not isinstance(valor, str):
        return valor
    texto_modificado = valor
    for buscar, reemplazo in REEMPLAZOS.items():
        patron = re.compile(re.escape(buscar), re.IGNORECASE)
        texto_modificado = patron.sub(reemplazo, texto_modificado)
    return texto_modificado

def hoja_a_dataframe(hoja):
    data = []
    for fila in hoja.iter_rows(values_only=True):
        data.append(fila)
    return pd.DataFrame(data)

def generar_pdf_desde_df(df):
    buffer = io.BytesIO()
    with PdfPages(buffer) as pdf:
        fig, ax = plt.subplots(figsize=(11.7, 8.3))  # A4 horizontal
        ax.axis('off')
        table = ax.table(
            cellText=df.values,
            colLabels=df.columns if df.columns[0] is not None else None,
            cellLoc='center',
            loc='center'
        )
        table.auto_set_font_size(False)
        table.set_fontsize(8)
        table.scale(1.2, 1.2)
        pdf.savefig(fig, bbox_inches='tight')
        plt.close(fig)
    buffer.seek(0)
    return buffer

if uploaded_file:
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded_file.read())
            tmp_path = tmp.name

        wb = load_workbook(tmp_path)
        hojas = wb.sheetnames
        hoja_seleccionada = st.selectbox("📄 Selecciona la hoja a adaptar", hojas)
        ws = wb[hoja_seleccionada]

        st.write("🔍 Vista previa original (primeras filas):")
        vista_previa = [[cell.value for cell in row] for row in ws.iter_rows(min_row=1, max_row=10)]
        st.dataframe(vista_previa)

        # Aplicar cambios
        for fila in ws.iter_rows():
            for celda in fila:
                if isinstance(celda.value, str):
                    celda.value = adaptar_valor(celda.value)

        # Guardar Excel adaptado
        output_excel = io.BytesIO()
        wb.save(output_excel)
        output_excel.seek(0)

        # Crear PDF
        df_pdf = hoja_a_dataframe(ws)
        pdf_buffer = generar_pdf_desde_df(df_pdf)

        # Botones de descarga
        st.success("✅ Menú adaptado con éxito.")

        st.download_button(
            label="📥 Descargar Excel adaptado",
            data=output_excel,
            file_name="menu_adaptado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.download_button(
            label="📄 Descargar como PDF",
            data=pdf_buffer,
            file_name="menu_adaptado.pdf",
            mime="application/pdf"
        )

    except Exception as e:
        st.error(f"❌ Error al procesar el archivo: {e}")




